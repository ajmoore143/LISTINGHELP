"""
Amazon flat-file auto-fill.

Fills the "Template" sheet of an Amazon category flat file (.xlsm) with the
content the app already produced (title, bullets, description, backend search
terms) plus product/business fields supplied by the worker. Anything left
empty is skipped so the worker can finish it by hand in Excel before upload.

The workbook's macros, data validations, dropdowns, and prefilled preference
defaults are preserved (openpyxl keep_vba=True; only data cells are written).
"""

from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple
import re

from openpyxl import load_workbook

MARKETPLACE_ID = "ATVPDKIKX0DER"  # amazon.com (US)
MP = MARKETPLACE_ID
LANG = "en_US"
TEMPLATE_SHEET = "Template"

# Fallback structural rows if the settings blob can't be parsed.
DEFAULT_LABEL_ROW = 4
DEFAULT_ATTRIBUTE_ROW = 5
DEFAULT_DATA_ROW = 8


# --- exact Amazon technical attribute names -------------------------------
# Only columns whose attribute name (row 5) matches one of these get written.
# Templates differ between FERTILIZER and SOIL; a key that does not exist as a
# column in the loaded template is silently skipped.

def _attr(logical: str, n: int = 1) -> str:
    m = {
        "item_name": f"item_name[marketplace_id={MP}][language_tag={LANG}]#1.value",
        "product_description": f"product_description[marketplace_id={MP}][language_tag={LANG}]#1.value",
        "brand": f"brand[marketplace_id={MP}][language_tag={LANG}]#1.value",
        "manufacturer": f"manufacturer[marketplace_id={MP}][language_tag={LANG}]#1.value",
        "country_of_origin": f"country_of_origin[marketplace_id={MP}]#1.value",
        "item_type_keyword": f"item_type_keyword[marketplace_id={MP}]#1.value",
        "product_type": "product_type#1.value",
        "parentage_level": f"parentage_level[marketplace_id={MP}]#1.value",
        "parent_sku": f"child_parent_sku_relationship[marketplace_id={MP}]#1.parent_sku",
        "variation_theme_name": "variation_theme#1.name",
        "size": f"size[marketplace_id={MP}][language_tag={LANG}]#1.value",
        "condition_type": f"condition_type[marketplace_id={MP}]#1.value",
        "dg_regulation": f"supplier_declared_dg_hz_regulation[marketplace_id={MP}]#1.value",
        "pesticide_status": f"pesticide_marking[marketplace_id={MP}]#1.registration_status",
        "shipping_group": f"merchant_shipping_group[marketplace_id={MP}]#1.value",
        "product_id_type": "amzn1.volt.ca.product_id_type",
        "product_id_value": "amzn1.volt.ca.product_id_value",
        "item_form": f"item_form[marketplace_id={MP}][language_tag={LANG}]#1.value",
        "ingredients": f"ingredients[marketplace_id={MP}][language_tag={LANG}]#1.value",
        "intended_use": f"intended_use[marketplace_id={MP}][language_tag={LANG}]#1.value",
        "contains_liquid": f"contains_liquid_contents[marketplace_id={MP}]#1.value",
        "prop65": f"california_proposition_65[marketplace_id={MP}]#1.chemical_names#1",
        "item_weight_value": f"item_weight[marketplace_id={MP}]#1.value",
        "item_weight_unit": f"item_weight[marketplace_id={MP}]#1.unit",
        "pkg_length": f"item_package_dimensions[marketplace_id={MP}]#1.length.value",
        "pkg_length_unit": f"item_package_dimensions[marketplace_id={MP}]#1.length.unit",
        "pkg_width": f"item_package_dimensions[marketplace_id={MP}]#1.width.value",
        "pkg_width_unit": f"item_package_dimensions[marketplace_id={MP}]#1.width.unit",
        "pkg_height": f"item_package_dimensions[marketplace_id={MP}]#1.height.value",
        "pkg_height_unit": f"item_package_dimensions[marketplace_id={MP}]#1.height.unit",
        "pkg_weight_value": f"item_package_weight[marketplace_id={MP}]#1.value",
        "pkg_weight_unit": f"item_package_weight[marketplace_id={MP}]#1.unit",
        "item_length": f"item_dimensions[marketplace_id={MP}]#1.length.value",
        "item_length_unit": f"item_dimensions[marketplace_id={MP}]#1.length.unit",
        "item_width": f"item_dimensions[marketplace_id={MP}]#1.width.value",
        "item_width_unit": f"item_dimensions[marketplace_id={MP}]#1.width.unit",
        "item_height": f"item_dimensions[marketplace_id={MP}]#1.height.value",
        "item_height_unit": f"item_dimensions[marketplace_id={MP}]#1.height.unit",
        "unit_count_value": f"unit_count[marketplace_id={MP}]#1.value",
        "unit_count_type": f"unit_count[marketplace_id={MP}]#1.type[language_tag={LANG}].value",
        "number_of_items": f"number_of_items[marketplace_id={MP}]#1.value",
        "number_of_packs": f"number_of_packs[marketplace_id={MP}]#1.value",
        "list_price_value": f"list_price[marketplace_id={MP}]#1.value",
        "list_price_currency": f"list_price[marketplace_id={MP}]#1.currency",
        "sku": "contribution_sku#1.value",
        "record_action": "::record_action",
        "fulfillment_channel_code": "fulfillment_availability#1.fulfillment_channel_code",
        "fulfillment_quantity": "fulfillment_availability#1.quantity",
    }
    if logical == "bullet_point":
        return f"bullet_point[marketplace_id={MP}][language_tag={LANG}]#{n}.value"
    if logical == "generic_keyword":
        return f"generic_keyword[marketplace_id={MP}][language_tag={LANG}]#{n}.value"
    return m[logical]


# Hard-required fields per category, used only to warn the worker which
# mandatory cells came out blank.
_REQUIRED = {
    "fertilizer": ["sku", "product_type", "item_name", "brand", "product_id_type",
                   "item_type_keyword", "product_description", "country_of_origin",
                   "dg_regulation"],
    "soil": ["sku", "item_name", "brand", "product_id_type", "item_type_keyword",
             "product_description", "country_of_origin", "dg_regulation"],
}


def _parse_structure(ws) -> Tuple[int, int, int]:
    """Read labelRow / attributeRow / dataRow from the row-1 settings blob."""
    label_row, attr_row, data_row = DEFAULT_LABEL_ROW, DEFAULT_ATTRIBUTE_ROW, DEFAULT_DATA_ROW
    try:
        blob = ws.cell(row=1, column=1).value or ""
        for key, default in (("labelRow", "label_row"), ("attributeRow", "attr_row"), ("dataRow", "data_row")):
            m = re.search(rf"{key}=(\d+)", str(blob))
            if m:
                val = int(m.group(1))
                if key == "labelRow":
                    label_row = val
                elif key == "attributeRow":
                    attr_row = val
                else:
                    data_row = val
    except Exception:
        pass
    return label_row, attr_row, data_row


def _column_map(ws, attr_row: int) -> Dict[str, int]:
    """attribute name (row 5 text) -> 1-based column index."""
    cols: Dict[str, int] = {}
    for idx, cell in enumerate(ws[attr_row], start=1):
        v = cell.value
        if v:
            cols[str(v).strip()] = idx
    return cols


def _is_blank(v: Any) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def build_flat_file(
    template_bytes: bytes,
    *,
    listing: Dict[str, Any],
    backend_search_terms: str,
    shared: Dict[str, Any],
    offers: List[Dict[str, Any]],
    product_kind: str = "fertilizer",
) -> Tuple[bytes, Dict[str, Any]]:
    """
    Write generated content + worker-supplied fields into the Template sheet.

    listing: {"titles": [...], "bullets": [...], "description": str}
    backend_search_terms: single backend keyword string (goes to generic_keyword#1)
    shared: logical_key -> value, applied identically to every offer row
    offers: list of {"sku","fulfillment_channel_code","quantity"}; one row each
    product_kind: "fertilizer" or "soil" (controls which required fields we check)

    Returns (xlsm_bytes, report). Empty values are skipped (left blank).
    """
    wb = load_workbook(BytesIO(template_bytes), keep_vba=True)
    if TEMPLATE_SHEET not in wb.sheetnames:
        raise ValueError(f"Sheet '{TEMPLATE_SHEET}' not found in template.")
    ws = wb[TEMPLATE_SHEET]

    _, attr_row, data_row = _parse_structure(ws)
    colmap = _column_map(ws, attr_row)

    # Capture the template's prefilled defaults on the first data row so every
    # offer row we add carries the same Amazon preference defaults.
    defaults = {c: ws.cell(row=data_row, column=c).value
                for c in range(1, ws.max_column + 1)
                if not _is_blank(ws.cell(row=data_row, column=c).value)}

    # --- assemble the shared (identical across rows) value set ---
    titles = listing.get("titles") or []
    bullets = listing.get("bullets") or []
    description = listing.get("description") or ""

    shared_values: Dict[str, Any] = {}

    def put(logical: str, value: Any, n: int = 1):
        if _is_blank(value):
            return
        attr = _attr(logical, n)
        if attr in colmap:
            shared_values[attr] = value

    if titles:
        put("item_name", titles[0])
    if description:
        put("product_description", description)
    for i, b in enumerate(bullets[:5], start=1):
        put("bullet_point", b, n=i)
    if backend_search_terms:
        put("generic_keyword", backend_search_terms, n=1)  # whole 249-byte string in #1

    # business/product fields supplied by the worker (all optional)
    for key, val in (shared or {}).items():
        if _is_blank(val):
            continue
        try:
            attr = _attr(key)
        except KeyError:
            continue
        if attr in colmap:
            shared_values[attr] = val
    # auto currency if a price was given and the column exists
    if not _is_blank(shared.get("list_price_value")) and _attr("list_price_currency") in colmap \
            and _attr("list_price_currency") not in shared_values:
        shared_values[_attr("list_price_currency")] = "USD"

    # --- write one row per offer (at least one row even with no offers) ---
    if not offers:
        offers = [{}]

    for r_off, offer in enumerate(offers):
        row = data_row + r_off
        # 1) lay down template defaults so every row is consistent
        for c, v in defaults.items():
            ws.cell(row=row, column=c, value=v)
        # 2) shared generated + business values
        for attr, val in shared_values.items():
            ws.cell(row=row, column=colmap[attr], value=val)
        # 3) per-offer values
        for key in ("sku", "fulfillment_channel_code", "fulfillment_quantity"):
            val = offer.get(key)
            if _is_blank(val):
                continue
            attr = _attr(key)
            if attr in colmap:
                ws.cell(row=row, column=colmap[attr], value=val)

    out = BytesIO()
    wb.save(out)
    wb.close()

    # --- report: what got filled, which required fields are blank ---
    required_logicals = _REQUIRED.get(product_kind, _REQUIRED["fertilizer"])
    first_row_has_sku = bool(offers and not _is_blank(offers[0].get("sku")))
    blank_required = []
    for logical in required_logicals:
        try:
            attr = _attr(logical)
        except KeyError:
            continue
        if attr not in colmap:
            continue
        if logical == "sku":
            if not first_row_has_sku:
                blank_required.append(logical)
        elif attr not in shared_values:
            blank_required.append(logical)

    report = {
        "rows_written": len(offers),
        "data_row": data_row,
        "fields_written": sorted({a.split("[")[0].split("#")[0] for a in shared_values}),
        "blank_required": blank_required,
        "template_columns": len(colmap),
    }
    return out.getvalue(), report


# --- shared descriptive fields copied onto parent + every child row ---
_DESCRIPTIVE_KEYS = [
    "brand", "manufacturer", "country_of_origin", "item_type_keyword",
    "condition_type", "dg_regulation", "pesticide_status", "shipping_group",
    "product_id_type", "item_form", "ingredients", "intended_use",
    "contains_liquid", "prop65",
]
# size-level keys that may be overridden per size (fall back to `shared`) ---
_SIZE_LEVEL_KEYS = [
    "product_id_value", "list_price_value", "list_price_currency",
    "item_weight_value", "item_weight_unit",
    "pkg_length", "pkg_length_unit", "pkg_width", "pkg_width_unit",
    "pkg_height", "pkg_height_unit", "pkg_weight_value", "pkg_weight_unit",
    "item_length", "item_length_unit", "item_width", "item_width_unit",
    "item_height", "item_height_unit", "unit_count_value", "unit_count_type",
    "number_of_items", "number_of_packs",
]


def build_family_flat_file(
    template_bytes: bytes,
    *,
    listing: Dict[str, Any],
    backend_search_terms: str,
    shared: Dict[str, Any],
    parent_sku: str,
    sizes: List[Dict[str, Any]],
    variation_theme: str = "Size",
    product_kind: str = "fertilizer",
    append_size_to_title: bool = True,
) -> Tuple[bytes, Dict[str, Any]]:
    """
    Build a parent-child variation flat file for a product family.

    One parent row + child rows. Each entry in `sizes` produces up to two child
    rows (one FBM, one FBA) that share the same `size` value and UPC, so Amazon
    links each FBM/FBA pair as two offers on one child ASIN, while different
    sizes become distinct child ASINs under the parent.

    sizes: list of dicts, e.g.
        {
          "size": "1/2 Quart",
          "product_id_value": "860...001",     # UPC for this size (both channels)
          "list_price_value": 14.99,
          "item_weight_value": 1.0, "item_weight_unit": "Pounds",
          "pkg_length": 6, ...,                 # any _SIZE_LEVEL_KEYS override
          "fbm_sku": "GW-FERT-QT-FBM", "fbm_qty": 50,
          "fba_sku": "GW-FERT-QT-FBA", "fba_qty": None,
        }
    Empty values are skipped. A size with only one channel SKU yields one child.
    """
    wb = load_workbook(BytesIO(template_bytes), keep_vba=True)
    if TEMPLATE_SHEET not in wb.sheetnames:
        raise ValueError(f"Sheet '{TEMPLATE_SHEET}' not found in template.")
    ws = wb[TEMPLATE_SHEET]
    _, attr_row, data_row = _parse_structure(ws)
    colmap = _column_map(ws, attr_row)

    defaults = {c: ws.cell(row=data_row, column=c).value
                for c in range(1, ws.max_column + 1)
                if not _is_blank(ws.cell(row=data_row, column=c).value)}

    titles = listing.get("titles") or []
    bullets = listing.get("bullets") or []
    description = listing.get("description") or ""
    family_title = titles[0] if titles else ""

    def col_for(logical: str, n: int = 1) -> Optional[int]:
        try:
            attr = _attr(logical, n)
        except KeyError:
            return None
        return colmap.get(attr)

    def write(row: int, logical: str, value: Any, n: int = 1):
        if _is_blank(value):
            return
        c = col_for(logical, n)
        if c:
            ws.cell(row=row, column=c, value=value)

    def write_descriptive(row: int):
        """Listing content + shared descriptive fields, identical on every row."""
        if description:
            write(row, "product_description", description)
        for i, b in enumerate(bullets[:5], start=1):
            write(row, "bullet_point", b, n=i)
        if backend_search_terms:
            write(row, "generic_keyword", backend_search_terms, n=1)
        write(row, "product_type", (shared.get("product_type") or product_kind.upper()))
        for key in _DESCRIPTIVE_KEYS:
            write(row, key, shared.get(key))

    rows_written = 0
    child_count = 0
    row = data_row

    # ---- parent row ----
    for c, v in defaults.items():
        ws.cell(row=row, column=c, value=v)
    write_descriptive(row)
    write(row, "item_name", family_title)
    write(row, "parentage_level", "Parent")
    write(row, "variation_theme_name", variation_theme)
    write(row, "sku", parent_sku)
    rows_written += 1

    # ---- child rows ----
    def size_val(size: Dict[str, Any], key: str):
        v = size.get(key)
        return v if not _is_blank(v) else shared.get(key)

    for size in sizes:
        size_label = size.get("size")
        child_title = family_title
        if append_size_to_title and not _is_blank(size_label) and family_title:
            child_title = f"{family_title} ({size_label})"[:200]

        offers = [
            ("DEFAULT", size.get("fbm_sku"), size.get("fbm_qty")),
            ("AMAZON_NA", size.get("fba_sku"), size.get("fba_qty")),
        ]
        for channel, sku, qty in offers:
            if _is_blank(sku):
                continue
            row += 1
            for c, v in defaults.items():
                ws.cell(row=row, column=c, value=v)
            write_descriptive(row)
            write(row, "item_name", child_title)
            write(row, "parentage_level", "Child")
            write(row, "parent_sku", parent_sku)
            write(row, "variation_theme_name", variation_theme)
            write(row, "size", size_label)
            write(row, "sku", sku)
            write(row, "fulfillment_channel_code", channel)
            write(row, "fulfillment_quantity", qty)
            for key in _SIZE_LEVEL_KEYS:
                write(row, key, size_val(size, key))
            # currency auto if a price exists
            if not _is_blank(size_val(size, "list_price_value")) \
                    and col_for("list_price_currency") and _is_blank(size.get("list_price_currency")) \
                    and _is_blank(shared.get("list_price_currency")):
                write(row, "list_price_currency", "USD")
            rows_written += 1
            child_count += 1

    out = BytesIO()
    wb.save(out)
    wb.close()

    report = {
        "rows_written": rows_written,
        "parent_sku": parent_sku,
        "children": child_count,
        "sizes": len(sizes),
        "variation_theme": variation_theme,
        "data_row": data_row,
        "template_columns": len(colmap),
    }
    return out.getvalue(), report
